# -*- coding: utf-8 -*-
"""
main.py - Backend FastAPI cho ứng dụng Giao Ban HTV
=====================================================
Quản lý cuộc họp giao ban hằng ngày, chỉ đạo TGĐ, sự kiện.
Dùng SSO + Local Auth.
"""
import os
import time
import io
import datetime
from fastapi import FastAPI, Request, HTTPException, Header
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field
from typing import Optional, List, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import db_service

try:
    from htv_sso_fastapi import (
        init_sso, get_sso_user, sso_exception_handler,
        _redirect_to_login, _extract_sso_department, _extract_sso_role,
    )
except ImportError:
    from htv_sso_fastapi import init_sso, get_sso_user, sso_exception_handler, _redirect_to_login

    def _extract_sso_department(p):
        return ""

    def _extract_sso_role(p):
        return "nhan_vien"


app = FastAPI(title="Giao Ban HTV")

# Cấu hình HTV SSO (giống Văn phòng Đài, session cookie riêng)
SSO_SECRET_KEY = os.environ.get("SSO_SECRET_KEY", "HTV_SSO_SHARED_SECRET_DOI_KHI_DEPLOY_THAT_2026")
SSO_SERVER_URL = os.environ.get("SSO_SERVER_URL", "http://ttphtl.htv.com.vn")

init_sso(app, secret_key=SSO_SECRET_KEY, sso_server_url=SSO_SERVER_URL, verify_slo=False, session_cookie="session_giaoban")
app.add_exception_handler(_redirect_to_login, sso_exception_handler)

# Tạo thư mục static và templates nếu chưa có
os.makedirs("static/css", exist_ok=True)
os.makedirs("static/js", exist_ok=True)
os.makedirs("templates", exist_ok=True)

# Cấu hình static và templates
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


# ===================== AUTH & HELPERS =====================

from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from htv_sso_fastapi import _login_url


# ===================== KIỂM TRA QUYỀN (dùng vai_tro SSO trực tiếp) =====================
#
# Không ánh xạ vai_tro SSO → role nội bộ. App đọc vai_tro nguyên gốc từ SSO.
#
# Bảng vai_tro SSO hiện tại:
#   BanTGD       → Ban Tổng Giám Đốc (TGĐ và PTGĐ)
#   truong_ban   → Trưởng đơn vị (ban/trung tâm)
#   pho_ban      → Phó đơn vị
#   truong_phong → Trưởng phòng (dùng nội bộ TTPHTL)
#   Pho_phong    → Phó phòng (dùng nội bộ TTPHTL)
#   nhan_vien    → Nhân viên (chỉ xem)
#
# Cột 'role' SSO (admin/editor/user) dùng để xác định Admin toàn hệ thống.


def is_admin_user(user: dict) -> bool:
    """Admin = SSO role='admin' HOẶC local DB role='Admin' HOẶC username='admin'."""
    username = (user.get("username") or "").strip().lower()
    if username == "admin":
        return True
    sso_role_col = (user.get("sso_role") or "").strip().lower()
    local_role = (user.get("role") or "").strip()
    vai_tro = (user.get("vai_tro") or "").strip()
    return sso_role_col == "admin" or local_role == "Admin" or vai_tro == "Admin"


def is_bantgd_user(user: dict) -> bool:
    """Ban TGĐ = SSO vai_tro='BanTGD' HOẶC local DB role='BanTGD' HOẶC user Ban TGĐ."""
    vai_tro = (user.get("vai_tro") or "").strip().lower()
    local_role = (user.get("role") or "").strip().lower()
    username = (user.get("username") or "").strip().lower()
    return (
        vai_tro in ["bantgd", "ban_tgd", "tgd"]
        or local_role in ["bantgd", "ban_tgd", "tgd"]
        or "tgd" in username
        or "caoanhminh" in username
        or "diepbuuchi" in username
    )


def is_ban_phu_trach(user: dict) -> bool:
    """Ban Phụ Trách = truong_ban / pho_ban / truong_phong / pho_phong.
    
    Tương đương 'trưởng/phó đơn vị' — có quyền nhập báo cáo của ban mình.
    Local login dùng role='BPT' cũng được tính.
    """
    vai_tro = (user.get("vai_tro") or "").strip().lower()
    local_role = (user.get("role") or "").strip()
    return (
        vai_tro in ["truong_ban", "pho_ban", "truong_phong", "pho_phong"]
        or local_role == "BPT"
    )


def is_vpd_user(user: dict) -> bool:
    """Có quyền quản lý cuộc họp / chỉ đạo:
    Admin, BanTGD, hoặc Ban Phụ Trách của Văn Phòng Đài.
    """
    if is_admin_user(user) or is_bantgd_user(user):
        return True
    dept = (user.get("department") or "").strip().lower()
    is_vpd_dept = any(k in dept for k in ["văn phòng đài", "van phong dai", "vpd", "vpđ"])
    return is_ban_phu_trach(user) and is_vpd_dept


def can_edit_report(user: dict, report_dept: str) -> bool:
    """Kiểm tra quyền sửa báo cáo của một đơn vị.
    
    - Admin/BanTGD/VPĐ BPT: sửa tất cả
    - BPT của đơn vị: chỉ sửa báo cáo của đơn vị mình
    """
    if is_admin_user(user) or is_bantgd_user(user):
        return True
    if is_vpd_user(user):
        return True
    if is_ban_phu_trach(user):
        user_dept = (user.get("department") or "").strip().lower()
        rd = (report_dept or "").strip().lower()
        return user_dept == rd or user_dept in rd or rd in user_dept
    return False


def get_current_user(request: Request) -> dict:
    """Lấy thông tin user hiện tại từ SSO hoặc Local session.
    
    Với SSO: vai_tro lấy nguyên gốc (truong_ban, pho_ban, BanTGD...)
    Với Local login: role lấy từ DB (Admin, BPT, BanTGD, nhan_vien)
    """
    sso_user = get_sso_user(request)
    if not sso_user:
        return {"logged_in": False}

    username = sso_user.get("username", "").strip()
    is_local = sso_user.get("is_local", False)
    raw_payload = sso_user.get("raw_payload") or {}
    full_name = sso_user.get("full_name") or username

    # Nếu là tài khoản Admin local
    if username.lower() == "admin":
        return {
            "logged_in": True,
            "username": "admin",
            "full_name": "Quản trị viên",
            "vai_tro": "Admin",
            "sso_role": "admin",
            "role": "Admin",
            "department": "Văn phòng Đài",
            "is_local": True,
        }

    if is_local:
        db_user = db_service.get_user(username)
        role = (db_user.get("Role") if db_user else None) or sso_user.get("role") or "nhan_vien"
        dept = (db_user.get("Department") if db_user else None) or sso_user.get("department") or ""
        return {
            "logged_in": True,
            "username": username,
            "full_name": full_name,
            "vai_tro": role,
            "sso_role": "admin" if role.lower() == "admin" else "user",
            "role": role,
            "department": dept,
            "is_local": True,
        }

    # --- Xử lý SSO ---
    vai_tro = (sso_user.get("vai_tro") or "").strip()
    if vai_tro.lower() in ["user", "guest", ""]:
        vai_tro = "nhan_vien"

    sso_role_col = (sso_user.get("role") or "user").strip()
    ban = (sso_user.get("ban") or sso_user.get("department") or "").strip()
    if not ban and raw_payload:
        ban = _extract_sso_department(raw_payload)

    # Chỉ đồng bộ user SSO vào DB, không ghi đè user local
    if username and username.lower() != "admin":
        db_service.save_or_update_sso_user(username, vai_tro, ban, force_update=True)

    return {
        "logged_in": True,
        "username": username,
        "full_name": full_name,
        "vai_tro": vai_tro,       # vai_tro SSO nguyên gốc
        "sso_role": sso_role_col, # role SSO ('admin'/'editor'/'user')
        "role": sso_role_col,
        "department": ban,
    }


# ===================== PYDANTIC MODELS =====================

class LoginRequest(BaseModel):
    username: str
    password: str


class UserCreateRequest(BaseModel):
    username: str
    password: str
    role: str
    department: Optional[str] = None


class UserUpdateRequest(BaseModel):
    role: str
    department: Optional[str] = ""


class MeetingCreateRequest(BaseModel):
    meeting_date: str = Field(..., alias="meetingDate")
    start_time: str = Field("08:00", alias="startTime")
    end_time: Optional[str] = Field(None, alias="endTime")
    location: str = Field("Phòng họp Giao ban Đài Phát thanh và Truyền hình Thành phố", alias="location")
    chairman: Optional[str] = None
    chairman_title: Optional[str] = Field(None, alias="chairmanTitle")
    secretary: Optional[str] = None
    secretary_title: Optional[str] = Field(None, alias="secretaryTitle")
    attendees: Optional[str] = None

    class Config:
        populate_by_name = True


class MeetingUpdateRequest(BaseModel):
    meeting_date: Optional[str] = Field(None, alias="meetingDate")
    start_time: Optional[str] = Field(None, alias="startTime")
    end_time: Optional[str] = Field(None, alias="endTime")
    location: Optional[str] = None
    chairman: Optional[str] = None
    chairman_title: Optional[str] = Field(None, alias="chairmanTitle")
    secretary: Optional[str] = None
    secretary_title: Optional[str] = Field(None, alias="secretaryTitle")
    attendees: Optional[str] = None
    status: Optional[str] = None

    class Config:
        populate_by_name = True


class ReportCreateRequest(BaseModel):
    department: str
    category: str = "noi_dung"
    content: str = ""


class DirectiveCreateRequest(BaseModel):
    category: str = "ket_luan"
    content: str
    assigned_to: Optional[str] = Field(None, alias="assignedTo")
    deadline: Optional[str] = None
    priority: int = 0

    class Config:
        populate_by_name = True


class DirectiveUpdateRequest(BaseModel):
    content: Optional[str] = None
    assigned_to: Optional[str] = Field(None, alias="assignedTo")
    deadline: Optional[str] = None
    status: Optional[str] = None
    priority: Optional[int] = None
    category: Optional[str] = None
    directive_date: Optional[str] = Field(None, alias="directiveDate")

    class Config:
        populate_by_name = True


class StandaloneDirectiveCreateRequest(BaseModel):
    category: str = "y_kien_tgd"
    content: str
    assigned_to: Optional[str] = Field(None, alias="assignedTo")
    deadline: Optional[str] = None
    priority: int = 0
    directive_date: Optional[str] = Field(None, alias="directiveDate")

    class Config:
        populate_by_name = True


class EventCreateRequest(BaseModel):
    title: str
    description: Optional[str] = None
    event_date: str = Field(..., alias="eventDate")
    event_end_date: Optional[str] = Field(None, alias="eventEndDate")
    event_type: str = Field("tuan", alias="eventType")

    class Config:
        populate_by_name = True


class EventUpdateRequest(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    event_date: Optional[str] = Field(None, alias="eventDate")
    event_end_date: Optional[str] = Field(None, alias="eventEndDate")
    event_type: Optional[str] = Field(None, alias="eventType")

    class Config:
        populate_by_name = True


# ===================== ROUTES =====================

@app.get("/login-sso")
async def trigger_sso_login(request: Request):
    return RedirectResponse(url=_login_url(request), status_code=302)


@app.get("/logout-local")
async def local_logout(request: Request):
    request.session.clear()
    html_content = """
    <!DOCTYPE html>
    <html><head><title>Đang đăng xuất...</title></head>
    <body><script>window.close();setTimeout(function(){window.open('','_self','');window.close();},100);</script></body>
    </html>
    """
    return HTMLResponse(content=html_content)


@app.get("/", response_class=HTMLResponse)
async def get_index(request: Request):
    token = request.query_params.get("token")
    if token:
        return RedirectResponse(url=f"/api/auth/sso?token={token}", status_code=303)

    sso_user = get_sso_user(request)
    referer = request.headers.get("referer", "")

    if not sso_user and (
        ("10.1.1.215" in referer or "ttphtl.htv.com.vn" in referer)
        or request.query_params.get("sso") == "true"
    ):
        return RedirectResponse(url=_login_url(request), status_code=302)

    return templates.TemplateResponse(request=request, name="index.html")


@app.get("/api/me")
def get_current_user_info(request: Request):
    user = get_current_user(request)
    if not user.get("logged_in"):
        return {"logged_in": False, "user": None}
    return user


@app.post("/api/login")
def login(request: Request, req: LoginRequest):
    user = db_service.get_user(req.username)
    if not user or user["Password"] != req.password:
        raise HTTPException(status_code=400, detail="Tài khoản hoặc mật khẩu không chính xác.")

    user_role = user["Role"] or "Admin"
    request.session["sso_user"] = {
        "username": user["Username"],
        "full_name": user["Username"],
        "role": user_role,
        "vai_tro": user_role,
        "department": user.get("Department") or "Văn phòng Đài",
        "is_local": True,
        "sver": 1,
    }
    request.session["_sso_last_check"] = time.time()

    return {
        "username": user["Username"],
        "role": user_role,
        "vai_tro": user_role,
        "department": user.get("Department") or "Văn phòng Đài",
    }


# ===================== USER MANAGEMENT =====================

@app.get("/api/users/new-count")
def get_new_users_count(request: Request):
    user = get_current_user(request)
    if not is_admin_user(user):
        return {"count": 0}
    return {"count": db_service.get_new_user_count()}


@app.get("/api/users")
def get_users(request: Request):
    user = get_current_user(request)
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập.")
    return db_service.get_all_users()


@app.post("/api/users")
def add_user(req: UserCreateRequest, request: Request):
    user = get_current_user(request)
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập.")
    existing = db_service.get_user(req.username)
    if existing:
        raise HTTPException(status_code=400, detail="Tên đăng nhập đã tồn tại.")
    db_service.create_user(req.username, req.password, req.role, req.department)
    return {"success": True, "message": "Tạo tài khoản thành công!"}


@app.put("/api/users/{username}")
def update_user_info(username: str, req: UserUpdateRequest, request: Request):
    user = get_current_user(request)
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập.")
    db_service.save_or_update_sso_user(username, req.role, req.department, force_update=True)
    return {"success": True, "message": f"Đã cập nhật tài khoản '{username}'!"}


@app.delete("/api/users/{username}")
def delete_user(username: str, request: Request):
    user = get_current_user(request)
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền truy cập.")
    if user.get("username", "").strip().lower() == username.strip().lower():
        raise HTTPException(status_code=400, detail="Bạn không thể tự xóa tài khoản của chính mình.")
    success = db_service.delete_user(username)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng.")
    return {"success": True, "message": "Xóa tài khoản thành công!"}


# ===================== MEETINGS =====================

def parse_date(date_str: str) -> str:
    """Chuyển DD-MM-YYYY hoặc YYYY-MM-DD sang YYYY-MM-DD."""
    if not date_str:
        return date_str
    try:
        dt = datetime.datetime.strptime(date_str, "%d-%m-%Y")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        try:
            datetime.datetime.strptime(date_str, "%Y-%m-%d")
            return date_str
        except ValueError:
            raise HTTPException(status_code=400, detail="Định dạng ngày không hợp lệ.")


@app.get("/api/meetings")
def api_get_meetings(
    request: Request,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    status: Optional[str] = None,
):
    db_start = parse_date(start_date) if start_date else None
    db_end = parse_date(end_date) if end_date else None
    meetings = db_service.get_meetings(start_date=db_start, end_date=db_end, status=status)
    return meetings


@app.post("/api/meetings")
def api_create_meeting(req: MeetingCreateRequest, request: Request):
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Chỉ BPT Văn phòng Đài hoặc Admin mới có quyền tạo cuộc họp.")

    db_date = parse_date(req.meeting_date)
    meeting_id = db_service.create_meeting(
        meeting_date=db_date,
        start_time=req.start_time,
        end_time=req.end_time,
        location=req.location,
        chairman=req.chairman,
        chairman_title=req.chairman_title,
        secretary=req.secretary,
        secretary_title=req.secretary_title,
        attendees=req.attendees,
        status="Draft",
        created_by=user.get("username"),
    )
    return {"success": True, "message": "Tạo cuộc họp thành công!", "meeting_id": meeting_id}


@app.get("/api/meetings/{meeting_id}")
def api_get_meeting(meeting_id: int):
    meeting = db_service.get_meeting(meeting_id)
    if not meeting:
        raise HTTPException(status_code=404, detail="Không tìm thấy cuộc họp.")
    return meeting


@app.put("/api/meetings/{meeting_id}")
def api_update_meeting(meeting_id: int, req: MeetingUpdateRequest, request: Request):
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Chỉ BPT Văn phòng Đài hoặc Admin mới có quyền sửa cuộc họp.")

    update_data = {}
    if req.meeting_date:
        update_data["MeetingDate"] = parse_date(req.meeting_date)
    if req.start_time is not None:
        update_data["StartTime"] = req.start_time
    if req.end_time is not None:
        update_data["EndTime"] = req.end_time
    if req.location is not None:
        update_data["Location"] = req.location
    if req.chairman is not None:
        update_data["Chairman"] = req.chairman
    if req.chairman_title is not None:
        update_data["ChairmanTitle"] = req.chairman_title
    if req.secretary is not None:
        update_data["Secretary"] = req.secretary
    if req.secretary_title is not None:
        update_data["SecretaryTitle"] = req.secretary_title
    if req.attendees is not None:
        update_data["Attendees"] = req.attendees
    if req.status is not None:
        update_data["Status"] = req.status

    success = db_service.update_meeting(meeting_id, **update_data)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy cuộc họp.")
    return {"success": True, "message": "Cập nhật cuộc họp thành công!"}


@app.delete("/api/meetings/{meeting_id}")
def api_delete_meeting(meeting_id: int, request: Request):
    user = get_current_user(request)
    if not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Chỉ Admin mới có quyền xóa cuộc họp.")
    success = db_service.delete_meeting(meeting_id)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy cuộc họp.")
    return {"success": True, "message": "Xóa cuộc họp thành công!"}


# ===================== REPORTS =====================

@app.get("/api/meetings/{meeting_id}/reports")
def api_get_reports(meeting_id: int, category: Optional[str] = None):
    return db_service.get_reports(meeting_id, category=category)


@app.post("/api/meetings/{meeting_id}/reports")
def api_create_report(meeting_id: int, req: ReportCreateRequest, request: Request):
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")

    if not can_edit_report(user, req.department):
        raise HTTPException(status_code=403, detail=f"Bạn không có quyền nhập báo cáo cho '{req.department}'.")

    meeting = db_service.get_meeting(meeting_id)
    if not meeting:
        raise HTTPException(status_code=404, detail="Không tìm thấy cuộc họp.")

    # BPT của ban khác không được nhập báo cáo khi cuộc họp đã Published
    if meeting.get("Status") == "Published" and not is_vpd_user(user) and not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Cuộc họp đã được công bố, không thể thêm báo cáo.")

    report_id = db_service.create_report(
        meeting_id=meeting_id,
        department=req.department,
        category=req.category,
        content=req.content,
        created_by=user.get("username"),
    )
    return {"success": True, "message": "Lưu báo cáo thành công!", "report_id": report_id}


@app.put("/api/meetings/{meeting_id}/reports/{report_id}")
def api_update_report(meeting_id: int, report_id: int, req: ReportCreateRequest, request: Request):
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")

    report = db_service.get_report(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Không tìm thấy báo cáo.")

    if not can_edit_report(user, report["Department"]):
        raise HTTPException(status_code=403, detail="Bạn không có quyền sửa báo cáo này.")

    # BPT của ban khác không được sửa báo cáo khi cuộc họp đã Published
    meeting = db_service.get_meeting(meeting_id)
    if meeting and meeting.get("Status") == "Published" and not is_vpd_user(user) and not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Cuộc họp đã được công bố, không thể sửa báo cáo.")

    success = db_service.update_report(report_id, req.content, created_by=user.get("username"))
    if not success:
        raise HTTPException(status_code=404, detail="Không thể cập nhật báo cáo.")
    return {"success": True, "message": "Cập nhật báo cáo thành công!"}


@app.delete("/api/meetings/{meeting_id}/reports/{report_id}")
def api_delete_report(meeting_id: int, report_id: int, request: Request):
    user = get_current_user(request)
    # BPT của ban mình cũng được xóa nếu họp chưa Published
    report = db_service.get_report(report_id)
    if not report:
        raise HTTPException(status_code=404, detail="Không tìm thấy báo cáo.")
    if not can_edit_report(user, report["Department"]):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xóa báo cáo này.")
    meeting = db_service.get_meeting(meeting_id)
    if meeting and meeting.get("Status") == "Published" and not is_vpd_user(user) and not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Cuộc họp đã công bố, không thể xóa báo cáo.")
    success = db_service.delete_report(report_id)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy báo cáo.")
    return {"success": True, "message": "Xóa báo cáo thành công!"}


# ===================== DIRECTIVES =====================

@app.get("/api/directives")
def api_get_all_directives(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    department: Optional[str] = None,
    days: Optional[int] = None,
    mode: Optional[str] = "default",  # "default" (today+yesterday), "all", "days", "custom"
    category: Optional[str] = None,
):
    """API lấy danh sách chỉ đạo TGĐ có hỗ trợ lọc theo Ban và Ngày."""
    if mode == "all":
        return db_service.get_directives_filtered(
            department=department,
            category=category,
        )
    elif days and days > 0:
        return db_service.get_recent_directives(days=days, department=department)
    elif start_date or end_date:
        db_start = parse_date(start_date) if start_date else None
        db_end = parse_date(end_date) if end_date else None
        return db_service.get_directives_filtered(
            start_date=db_start,
            end_date=db_end,
            department=department,
            category=category,
        )
    else:
        # Mặc định: hôm nay và hôm qua (có fallback 7 ngày nếu chưa có dữ liệu)
        return db_service.get_recent_directives_2days(department=department)


@app.post("/api/directives")
def api_create_standalone_directive(req: StandaloneDirectiveCreateRequest, request: Request):
    """API tạo chỉ đạo ngoài cuộc họp (Ban Tổng Giám đốc)."""
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")
    if not is_bantgd_user(user) and not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Chỉ Ban Tổng Giám đốc hoặc Admin mới có quyền thêm chỉ đạo ngoài cuộc họp.")

    directive_date = req.directive_date or None
    if directive_date:
        directive_date = parse_date(directive_date)

    directive_id = db_service.create_standalone_directive(
        category=req.category,
        content=req.content,
        assigned_to=req.assigned_to,
        deadline=parse_date(req.deadline) if req.deadline else None,
        priority=req.priority,
        directive_date=directive_date,
        created_by=user.get("username"),
    )
    return {"success": True, "message": "Thêm chỉ đạo thành công!", "directive_id": directive_id}


@app.put("/api/directives/{directive_id}")
def api_update_standalone_directive(directive_id: int, req: DirectiveUpdateRequest, request: Request):
    """API sửa chỉ đạo ngoài cuộc họp."""
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")
    if not is_bantgd_user(user) and not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Không có quyền sửa chỉ đạo này.")

    update_data = {}
    if req.content is not None:
        update_data["Content"] = req.content
    if req.assigned_to is not None:
        update_data["AssignedTo"] = req.assigned_to
    if req.deadline is not None:
        update_data["Deadline"] = parse_date(req.deadline) if req.deadline else None
    if req.status is not None:
        update_data["Status"] = req.status
    if req.priority is not None:
        update_data["Priority"] = req.priority
    if req.category is not None:
        update_data["Category"] = req.category
    if req.directive_date is not None:
        update_data["DirectiveDate"] = parse_date(req.directive_date) if req.directive_date else None

    success = db_service.update_standalone_directive(directive_id, **update_data)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy chỉ đạo.")
    return {"success": True, "message": "Cập nhật chỉ đạo thành công!"}


@app.delete("/api/directives/{directive_id}")
def api_delete_standalone_directive(directive_id: int, request: Request):
    """API xóa chỉ đạo ngoài cuộc họp."""
    user = get_current_user(request)
    if not is_bantgd_user(user) and not is_admin_user(user):
        raise HTTPException(status_code=403, detail="Không có quyền xóa chỉ đạo này.")
    success = db_service.delete_directive(directive_id)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy chỉ đạo.")
    return {"success": True, "message": "Xóa chỉ đạo thành công!"}



@app.get("/api/directives/today")
def api_get_today_directives(department: Optional[str] = None):
    """Lấy chỉ đạo (mặc định hôm nay & hôm qua)."""
    return db_service.get_recent_directives_2days(department=department)


@app.get("/api/directives/recent")
def api_get_recent_directives(days: int = 7, department: Optional[str] = None):
    """Lấy chỉ đạo N ngày gần nhất."""
    return db_service.get_recent_directives(days=days, department=department)


@app.get("/api/meetings/{meeting_id}/directives")
def api_get_directives(meeting_id: int, category: Optional[str] = None):
    return db_service.get_directives(meeting_id=meeting_id, category=category)


@app.post("/api/meetings/{meeting_id}/directives")
def api_create_directive(meeting_id: int, req: DirectiveCreateRequest, request: Request):
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Chỉ BPT Văn phòng Đài hoặc Admin mới có quyền nhập chỉ đạo.")

    meeting = db_service.get_meeting(meeting_id)
    if not meeting:
        raise HTTPException(status_code=404, detail="Không tìm thấy cuộc họp.")

    directive_id = db_service.create_directive(
        meeting_id=meeting_id,
        category=req.category,
        content=req.content,
        assigned_to=req.assigned_to,
        deadline=req.deadline,
        priority=req.priority,
        created_by=user.get("username"),
    )
    return {"success": True, "message": "Thêm chỉ đạo thành công!", "directive_id": directive_id}


@app.put("/api/meetings/{meeting_id}/directives/{directive_id}")
def api_update_directive(meeting_id: int, directive_id: int, req: DirectiveUpdateRequest, request: Request):
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Chỉ BPT Văn phòng Đài hoặc Admin mới có quyền sửa chỉ đạo.")

    update_data = {}
    if req.content is not None:
        update_data["Content"] = req.content
    if req.assigned_to is not None:
        update_data["AssignedTo"] = req.assigned_to
    if req.deadline is not None:
        update_data["Deadline"] = req.deadline
    if req.status is not None:
        update_data["Status"] = req.status
    if req.priority is not None:
        update_data["Priority"] = req.priority
    if req.category is not None:
        update_data["Category"] = req.category

    success = db_service.update_directive(directive_id, **update_data)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy chỉ đạo.")
    return {"success": True, "message": "Cập nhật chỉ đạo thành công!"}


@app.delete("/api/meetings/{meeting_id}/directives/{directive_id}")
def api_delete_directive(meeting_id: int, directive_id: int, request: Request):
    user = get_current_user(request)
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xóa chỉ đạo.")
    success = db_service.delete_directive(directive_id)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy chỉ đạo.")
    return {"success": True, "message": "Xóa chỉ đạo thành công!"}


# ===================== EVENTS =====================

@app.get("/api/events")
def api_get_events(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    event_type: Optional[str] = None,
):
    db_start = parse_date(start_date) if start_date else None
    db_end = parse_date(end_date) if end_date else None
    return db_service.get_events(start_date=db_start, end_date=db_end, event_type=event_type)


@app.post("/api/events")
def api_create_event(req: EventCreateRequest, request: Request):
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Chỉ BPT Văn phòng Đài hoặc Admin mới có quyền thêm sự kiện.")

    event_id = db_service.create_event(
        title=req.title,
        event_date=parse_date(req.event_date),
        event_end_date=parse_date(req.event_end_date) if req.event_end_date else None,
        description=req.description,
        event_type=req.event_type,
        created_by=user.get("username"),
    )
    return {"success": True, "message": "Thêm sự kiện thành công!", "event_id": event_id}


@app.put("/api/events/{event_id}")
def api_update_event(event_id: int, req: EventUpdateRequest, request: Request):
    user = get_current_user(request)
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền sửa sự kiện.")

    update_data = {}
    if req.title is not None:
        update_data["Title"] = req.title
    if req.description is not None:
        update_data["Description"] = req.description
    if req.event_date is not None:
        update_data["EventDate"] = parse_date(req.event_date)
    if req.event_end_date is not None:
        update_data["EventEndDate"] = parse_date(req.event_end_date)
    if req.event_type is not None:
        update_data["EventType"] = req.event_type

    success = db_service.update_event(event_id, **update_data)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy sự kiện.")
    return {"success": True, "message": "Cập nhật sự kiện thành công!"}


@app.delete("/api/events/{event_id}")
def api_delete_event(event_id: int, request: Request):
    user = get_current_user(request)
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xóa sự kiện.")
    success = db_service.delete_event(event_id)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy sự kiện.")
    return {"success": True, "message": "Xóa sự kiện thành công!"}


# ===================== PROPAGANDA PLANS (Kế hoạch tuyên truyền) =====================

class PropagandaPlanCreateRequest(BaseModel):
    activity_name: str = Field(..., alias="activityName")
    organizer: Optional[str] = None
    executing_unit: Optional[str] = Field(None, alias="executingUnit")
    event_time: Optional[str] = Field(None, alias="eventTime")
    location: Optional[str] = None
    assigned_unit: Optional[str] = Field(None, alias="assignedUnit")
    cooperating_unit: Optional[str] = Field(None, alias="cooperatingUnit")
    notes: Optional[str] = None
    plan_date: str = Field(..., alias="planDate")
    plan_end_date: Optional[str] = Field(None, alias="planEndDate")

    class Config:
        populate_by_name = True


class PropagandaPlanUpdateRequest(BaseModel):
    activity_name: Optional[str] = Field(None, alias="activityName")
    organizer: Optional[str] = None
    executing_unit: Optional[str] = Field(None, alias="executingUnit")
    event_time: Optional[str] = Field(None, alias="eventTime")
    location: Optional[str] = None
    assigned_unit: Optional[str] = Field(None, alias="assignedUnit")
    cooperating_unit: Optional[str] = Field(None, alias="cooperatingUnit")
    notes: Optional[str] = None
    plan_date: Optional[str] = Field(None, alias="planDate")
    plan_end_date: Optional[str] = Field(None, alias="planEndDate")

    class Config:
        populate_by_name = True


@app.get("/api/propaganda-plans")
def api_get_propaganda_plans(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    upcoming: Optional[bool] = False,
    days: Optional[int] = 60,
):
    """Lấy danh sách kế hoạch tuyên truyền."""
    if upcoming:
        return db_service.get_upcoming_propaganda_plans(days=days)
    db_start = parse_date(start_date) if start_date else None
    db_end = parse_date(end_date) if end_date else None
    return db_service.get_propaganda_plans(start_date=db_start, end_date=db_end)


@app.get("/api/propaganda-plans/export")
def api_export_propaganda_plans(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """Xuất danh sách kế hoạch tuyên truyền ra file Excel (.xlsx)."""
    db_start = parse_date(start_date) if start_date else None
    db_end = parse_date(end_date) if end_date else None
    plans = db_service.get_propaganda_plans(start_date=db_start, end_date=db_end, limit=1000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KHTuyenTruyen"

    # Header title
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "KẾ HOẠCH TUYÊN TRUYỀN - ĐÀI PHÁT THANH VÀ TRUYỀN HÌNH TP.HCM (HTV)"
    title_cell.font = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Sub title / export time
    ws.merge_cells("A2:K2")
    sub_cell = ws["A2"]
    now_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    filter_desc = "Tất cả kế hoạch"
    if db_start and db_end:
        filter_desc = f"Thời gian: {db_start} đến {db_end}"
    elif db_start:
        filter_desc = f"Từ ngày: {db_start}"
    elif db_end:
        filter_desc = f"Đến ngày: {db_end}"
    sub_cell.value = f"{filter_desc} • Xuất ngày: {now_str}"
    sub_cell.font = Font(name="Arial", size=10, italic=True)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Headers table (row 3)
    headers = [
        "STT", "Tên hoạt động", "Ngày bắt đầu", "Ngày kết thúc", 
        "Thời gian", "Danh nghĩa tổ chức", "Đơn vị thực hiện", 
        "Địa điểm", "Phân công đơn vị HTV", "Đơn vị phối hợp", "Ghi chú"
    ]
    header_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7')
    )

    ws.row_dimensions[3].height = 26
    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = h_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    row_idx = 4
    for idx, p in enumerate(plans, 1):
        ws.row_dimensions[row_idx].height = 24
        p_date = p.get("PlanDate") or ""
        p_end = p.get("PlanEndDate") or ""

        row_values = [
            idx,
            p.get("ActivityName") or "",
            p_date,
            p_end,
            p.get("EventTime") or "",
            p.get("Organizer") or "",
            p.get("ExecutingUnit") or "",
            p.get("Location") or "",
            p.get("AssignedUnit") or "",
            p.get("CooperatingUnit") or "",
            p.get("Notes") or ""
        ]

        row_fill = PatternFill(start_color="FFFFFF" if idx % 2 == 1 else "FFF8F0", fill_type="solid")
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border
            cell.fill = row_fill
            if col_idx in [1, 3, 4]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row_idx += 1

    # Widths
    column_widths = {
        1: 6,   # STT
        2: 36,  # Tên hoạt động
        3: 14,  # Ngày bắt đầu
        4: 14,  # Ngày kết thúc
        5: 25,  # Thời gian
        6: 24,  # Danh nghĩa tổ chức
        7: 25,  # Đơn vị thực hiện
        8: 30,  # Địa điểm
        9: 25,  # Phân công đơn vị HTV
        10: 25, # Đơn vị phối hợp
        11: 28  # Ghi chú
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = "Ke_hoach_tuyen_truyen_HTV.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/propaganda-plans/{plan_id}")
def api_get_propaganda_plan(plan_id: int):
    plan = db_service.get_propaganda_plan(plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Không tìm thấy kế hoạch.")
    return plan


@app.post("/api/propaganda-plans")
def api_create_propaganda_plan(req: PropagandaPlanCreateRequest, request: Request):
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Chỉ BPT Văn phòng Đài hoặc Admin mới có quyền thêm kế hoạch tuyên truyền.")
    plan_id = db_service.create_propaganda_plan(
        activity_name=req.activity_name,
        organizer=req.organizer,
        executing_unit=req.executing_unit,
        event_time=req.event_time,
        location=req.location,
        assigned_unit=req.assigned_unit,
        cooperating_unit=req.cooperating_unit,
        notes=req.notes,
        plan_date=parse_date(req.plan_date),
        plan_end_date=parse_date(req.plan_end_date) if req.plan_end_date else None,
        created_by=user.get("username"),
    )
    return {"success": True, "message": "Thêm kế hoạch tuyên truyền thành công!", "plan_id": plan_id}


@app.put("/api/propaganda-plans/{plan_id}")
def api_update_propaganda_plan(plan_id: int, req: PropagandaPlanUpdateRequest, request: Request):
    user = get_current_user(request)
    if not user.get("logged_in"):
        raise HTTPException(status_code=401, detail="Vui lòng đăng nhập.")
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền sửa kế hoạch này.")
    update_data = {}
    if req.activity_name is not None:
        update_data["ActivityName"] = req.activity_name
    if req.organizer is not None:
        update_data["Organizer"] = req.organizer
    if req.executing_unit is not None:
        update_data["ExecutingUnit"] = req.executing_unit
    if req.event_time is not None:
        update_data["EventTime"] = req.event_time
    if req.location is not None:
        update_data["Location"] = req.location
    if req.assigned_unit is not None:
        update_data["AssignedUnit"] = req.assigned_unit
    if req.cooperating_unit is not None:
        update_data["CooperatingUnit"] = req.cooperating_unit
    if req.notes is not None:
        update_data["Notes"] = req.notes
    if req.plan_date is not None:
        update_data["PlanDate"] = parse_date(req.plan_date)
    if req.plan_end_date is not None:
        update_data["PlanEndDate"] = parse_date(req.plan_end_date) if req.plan_end_date else None
    success = db_service.update_propaganda_plan(plan_id, **update_data)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy kế hoạch.")
    return {"success": True, "message": "Cập nhật kế hoạch thành công!"}


@app.delete("/api/propaganda-plans/{plan_id}")
def api_delete_propaganda_plan(plan_id: int, request: Request):
    user = get_current_user(request)
    if not is_vpd_user(user):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xóa kế hoạch này.")
    success = db_service.delete_propaganda_plan(plan_id)
    if not success:
        raise HTTPException(status_code=404, detail="Không tìm thấy kế hoạch.")
    return {"success": True, "message": "Xóa kế hoạch thành công!"}


# ===================== HEARTBEAT =====================

@app.post("/api/heartbeat")
async def post_heartbeat():
    return {"status": "ok"}
